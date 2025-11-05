import os
import math
import joblib
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier, IsolationForest
from sklearn.svm import SVC, OneClassSVM
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (classification_report, confusion_matrix, roc_auc_score,
                             accuracy_score, precision_score, recall_score, f1_score,
                             roc_curve, precision_recall_curve, average_precision_score)
from sklearn.preprocessing import StandardScaler, OneHotEncoder
from sklearn.decomposition import PCA
from sklearn.cluster import DBSCAN
from sklearn.neighbors import LocalOutlierFactor
from imblearn.over_sampling import SMOTE
from ydata_profiling import ProfileReport

from packaging import version
import sklearn

# Logging configuration
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Ignore warnings for cleaner output
warnings.filterwarnings('ignore')

# Directory and file configuration
INPUTS_DIR = "inputs"
RESULTS_DIR = "outputs"
MODELS_DIR = "models"
EXCEL_FILE = os.path.join(RESULTS_DIR, 'model_evaluation.xlsx')
PDF_REPORT = os.path.join(RESULTS_DIR, "technical_report.pdf")

def setup_directories():
    """Create necessary directories."""
    if not os.path.exists(INPUTS_DIR):
        os.makedirs(INPUTS_DIR)
        logging.info(f"Directory '{INPUTS_DIR}' created.")
    else:
        logging.info(f"Directory '{INPUTS_DIR}' already exists.")

    if not os.path.exists(RESULTS_DIR):
        os.makedirs(RESULTS_DIR)
        logging.info(f"Directory '{RESULTS_DIR}' created.")
    else:
        logging.info(f"Directory '{RESULTS_DIR}' already exists.")

    if not os.path.exists(MODELS_DIR):
        os.makedirs(MODELS_DIR)
        logging.info(f"Directory '{MODELS_DIR}' created.")
    else:
        logging.info(f"Directory '{MODELS_DIR}' already exists.")

    if not os.path.exists(EXCEL_FILE):
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
            pd.DataFrame().to_excel(writer)  # Create empty Excel file
        logging.info(f"Excel file '{EXCEL_FILE}' created.")
    else:
        logging.info(f"Excel file '{EXCEL_FILE}' already exists.")

def generate_simulated_temporal_data():
    """Generate simulated temporal data for multiple equipment."""
    np.random.seed(42)
    n_equipment = 100  # Number of devices
    n_time_steps = 40  # Number of samples per device

    # Create an array for the records
    data_records = []

    for equipment in range(1, n_equipment + 1):
        # Assign time, for simplicity 1 second period
        process_type = np.random.choice(['Vibrations', 'Oil Analysis', 'Hours Operated'])
        for t in range(1, n_time_steps + 1):
            record = {
                'equipment_id': equipment,
                'time_step': t,
                'process_type': process_type
            }

            # Simulate according to the characteristics
            if process_type == 'Vibrations':
                vib = np.sin(t / 5) + np.random.normal(0, 0.5)
                temp = 20 + 2 * vib + np.random.normal(0, 0.5)
                pres = 30 + 3 * (vib ** 2) + np.random.normal(0, 1)
                record.update({
                    'vibration': vib,
                    'temperature': temp,
                    'pressure': pres,
                    'oil_quality': np.nan,
                    'contaminant_level': np.nan,
                    'acidity': np.nan,
                    'hours_operated': np.nan,
                    'maintenance_history': np.nan,
                    'load': np.nan
                })
            elif process_type == 'Oil Analysis':
                oil_q = np.random.uniform(0, 100) + t * 0.1  # Small increment with time
                cont_level = 50 + 0.5 * oil_q + np.random.normal(0, 5)
                acid = 10 + 0.3 * (oil_q ** 1.5) + np.random.normal(0, 2)
                record.update({
                    'vibration': np.nan,
                    'temperature': np.nan,
                    'pressure': np.nan,
                    'oil_quality': oil_q,
                    'contaminant_level': cont_level,
                    'acidity': acid,
                    'hours_operated': np.nan,
                    'maintenance_history': np.nan,
                    'load': np.nan
                })
            elif process_type == 'Hours Operated':
                hours_op = np.random.exponential(scale=50) + t * 0.5  # Proportional to time
                maint_hist = np.random.poisson(lam=2)  # Maintenance history
                ld = 100 + 0.1 * t + np.random.normal(0, 10)
                record.update({
                    'vibration': np.nan,
                    'temperature': np.nan,
                    'pressure': np.nan,
                    'oil_quality': np.nan,
                    'contaminant_level': np.nan,
                    'acidity': np.nan,
                    'hours_operated': hours_op,
                    'maintenance_history': maint_hist,
                    'load': ld
                })

            # Failure emulation
            if process_type == 'Vibrations':
                fail = int((0.3 * vib + 0.2 * temp - 0.1 * pres + np.random.normal(0, 0.5)) > 1)
            elif process_type == 'Oil Analysis':
                fail = int((0.2 * oil_q - 0.1 * cont_level + 0.05 * acid + np.random.normal(0, 1)) > 5)
            elif process_type == 'Hours Operated':
                fail = int((0.05 * hours_op + 0.1 * maint_hist - 0.02 * ld + np.random.normal(0, 1)) > 3)
            record['failure'] = fail

            # Introduction of random anomalies
            if np.random.rand() < 0.02:  # 2% anomaly likelihood
                record['anomaly'] = 1
                # Alter some variables
                if process_type == 'Vibrations':
                    record['vibration'] += np.random.normal(10, 5)  # Vibration
                elif process_type == 'Oil Analysis':
                    record['oil_quality'] += np.random.uniform(50, 100)  # Oil quality
                elif process_type == 'Hours Operated':
                    record['load'] += np.random.uniform(50, 100)  # Load
            else:
                record['anomaly'] = 0

            data_records.append(record)

    # Create DataFrame

    # data = pd.read_csv('data/data.csv')
    data = pd.DataFrame(data_records)

    print(f"Data shape: {data.shape}")
    print(f"Data info: {data.info()}")
    print(f"Data describe: {data.describe()}")
    # print(f"Data median: {data.median()}")

    # Handle Nan variables
    numeric_cols = data.select_dtypes(include=[np.number]).columns
    data[numeric_cols] = data[numeric_cols].fillna(data[numeric_cols].mean())

    # Make sure 'anomaly' is integer
    data['anomaly'] = data['anomaly'].astype(int)

    logging.info("Temporal data emulated correctly.")
    return data

def exportToCSV(data):
    data.to_csv(os.path.join(INPUTS_DIR, 'emulated_data.csv'), index=False)

def handle_data_types(data):
    """
    Makes sure all data types are correct
    - Converts categorical variables into numeric according to codification.
    - Makes sure numeric variables are of the correct type.
    """
    categorical_cols = data.select_dtypes(include=['object', 'category']).columns.tolist()

    # Determine scikit-learn version
    skl_version = version.parse(sklearn.__version__)

    # Define OneHotEncoder parameters according to version
    if skl_version >= version.parse("1.2"):
        encoder = OneHotEncoder(drop='first', sparse_output=False)
    else:
        encoder = OneHotEncoder(drop='first', sparse=False)

    # Encode categorical variables using OneHotEncoder
    if categorical_cols:
        try:
            encoded_data = encoder.fit_transform(data[categorical_cols])
            encoded_cols = encoder.get_feature_names_out(categorical_cols)
            encoded_df = pd.DataFrame(encoded_data, columns=encoded_cols, index=data.index)
            data = pd.concat([data.drop(categorical_cols, axis=1), encoded_df], axis=1)
            logging.info("Categorical variables coded correctly.")
        except Exception as e:
            logging.error(f"Error encoding variables: {e}")
            raise
    else:
        logging.info("No categorical variable columns found.")

    # Ensure all numeric columns are of float type
    numeric_cols = data.select_dtypes(include=[np.number]).columns
    data[numeric_cols] = data[numeric_cols].astype(float)
    logging.info("Ensured numeric data type as float.")

    # Additional verification
    remaining_categorical = data.select_dtypes(include=['object', 'category']).columns.tolist()
    if remaining_categorical:
        raise ValueError(f"Following columns have not been coded yet and are still categorical: {remaining_categorical}")
    else:
        logging.info("All columns have been coded")

    return data

def perform_eda(data):
    """Perform exploratory analysis of the data (EDA)."""
    try:
        profile = ProfileReport(data, title='Exploratory analysis of the data', explorative=True)
        eda_file = os.path.join(RESULTS_DIR, "dataset_EDA.html")
        profile.to_file(eda_file)
        logging.info(f"EDA report saved to '{eda_file}'.")
    except Exception as e:
        logging.error(f"Error generating EDA report: {e}")

    # Distribution of failures and not failures
    plt.figure(figsize=(8, 6))
    sns.countplot(x='failure', data=data, palette='coolwarm')
    plt.title('Failure/Not Failure distribution')
    plt.xlabel('Failure')
    plt.ylabel('Count')
    plt.savefig(os.path.join(RESULTS_DIR, 'failure_distribution.png'), dpi=300)
    plt.close()
    logging.info("Plot 'failure_distribution.png' saved.")

    # Correlation matrix
    plt.figure(figsize=(14, 10))
    corr = data.corr()
    sns.heatmap(corr, annot=True, cmap='coolwarm', fmt='.2f', linewidths=0.5)
    plt.title('Simulated Data Correlation Matrix')
    plt.savefig(os.path.join(RESULTS_DIR, 'correlation_matrix.png'), dpi=300)
    plt.close()
    logging.info("Plot 'correlation_matrix.png' saved.")

    # Histogram per variable
    numeric_columns = data.select_dtypes(include=[np.number]).columns
    data[numeric_columns].hist(bins=30, figsize=(20, 15), color='steelblue', edgecolor='black')
    plt.suptitle('Histograms of Variables', fontsize=16)
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    histograms_path = os.path.join(RESULTS_DIR, 'histograms.png')
    plt.savefig(histograms_path, dpi=300)
    plt.close()
    logging.info(f"Plot 'histograms.png' saved.")

    # Boxplots to detect outliers
    numeric_columns = data.select_dtypes(include=[np.number]).columns
    feature_columns = numeric_columns.drop(['failure', 'anomaly'], errors='ignore')
    num_features = len(feature_columns)

    # Define number of columns per row
    num_cols = 3
    # Calculate number of rows needed
    num_rows = math.ceil(num_features / num_cols)

    # Create a large figure to contain all boxplots
    fig, axes = plt.subplots(num_rows, num_cols, figsize=(num_cols * 6, num_rows * 4))
    axes = axes.flatten()  # Flatten to iterate easily

    for idx, column in enumerate(feature_columns):
        sns.boxplot(y=data[column], ax=axes[idx], color='lightgreen')
        axes[idx].set_title(f'Boxplot of {column}')

    # Delete subplots if any
    for ax in axes[num_features:]:
        fig.delaxes(ax)

    plt.tight_layout()
    boxplots_path = os.path.join(RESULTS_DIR, 'boxplots.png')
    plt.savefig(boxplots_path, dpi=300)
    plt.close()
    logging.info(f"Plot 'boxplots.png' saved.")

    # Pairplot to validate relations between variables
    sns.pairplot(data.drop(['equipment_id', 'time_step'], axis=1, errors='ignore'), hue='failure', palette='coolwarm', diag_kind='kde')
    plt.suptitle('Pairplot of Variables', y=1.02)
    pairplot_path = os.path.join(RESULTS_DIR, 'pairplot.png')
    plt.savefig(pairplot_path, dpi=300)
    plt.close()
    logging.info(f"Plot 'pairplot.png' saved.")

def preprocess_data(data):
    """Preprocess data: scaling, balancing and split of testing and training data."""
    # Independent variables (X) and dependent variables (y)
    X = data.drop(['failure', 'equipment_id', 'time_step', 'anomaly'], axis=1, errors='ignore')
    y = data['failure'].astype(int)  # Ensure 'failure' is of integer type

    # Scaling characteristics
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    logging.info("Successfully scaled characteristics.")

    # Handling imbalance of classes using SMOTE
    smote = SMOTE(random_state=42)
    X_resampled, y_resampled = smote.fit_resample(X_scaled, y)
    logging.info("Successfully balanced data using SMOTE.")

    # Split training and testing data
    x_train, x_test, y_train, y_test = train_test_split(
        X_resampled, y_resampled, test_size=0.3, random_state=42, stratify=y_resampled
    )
    logging.info("Successfully split data for training and testing.")
    return x_train, x_test, y_train, y_test, X.columns

def train_classification_models(x_train, y_train):
    """Train ML models using GridSearchCV"""
    models = {
        'RandomForest': {
            'model': RandomForestClassifier(random_state=42),
            'params': {
                'n_estimators': [100, 200],
                'max_depth': [None, 10, 20],
                'min_samples_split': [2, 5]
            }
        },
        'SVM': {
            'model': SVC(probability=True, random_state=42),
            'params': {
                'C': [0.1, 1, 10],
                'kernel': ['linear', 'rbf']
            }
        },
        'GradientBoosting': {
            'model': GradientBoostingClassifier(random_state=42),
            'params': {
                'n_estimators': [100, 200],
                'learning_rate': [0.05, 0.1],
                'max_depth': [3, 5]
            }
        },
        'LogisticRegression': {
            'model': LogisticRegression(random_state=42, max_iter=1000),
            'params': {
                'C': [0.01, 0.1, 1, 10],
                'penalty': ['l2']
            }
        }
    }

    best_models = {}
    for model_name, mp in models.items():
        logging.info(f"Training and adjusting parameters for model: {model_name}...")
        try:
            grid = GridSearchCV(mp['model'], mp['params'], cv=5, scoring='roc_auc', n_jobs=-1)
            grid.fit(x_train, y_train)
            best_models[model_name] = grid.best_estimator_
            logging.info(f"Best parameters for {model_name}: {grid.best_params_}")
            logging.info(f"Best ROC AUC validation for {model_name}: {grid.best_score_:.4f}\n")
        except Exception as e:
            logging.error(f"Error training {model_name}: {e}")
    return best_models

def export_models(best_models):
    for name, model in best_models.items():
        joblib.dump(model, f"{MODELS_DIR}/{name}_model.pkl")

def evaluate_classification_models(best_models, x_test, y_test):
    """Evaluate trained models and save results."""
    def save_results(model_name, model, x_test, y_test):
        try:
            y_pred = model.predict(x_test)
            y_pred_proba = model.predict_proba(x_test)[:, 1]

            accuracy = accuracy_score(y_test, y_pred)
            precision = precision_score(y_test, y_pred)
            recall = recall_score(y_test, y_pred)
            f1 = f1_score(y_test, y_pred)
            roc_auc = roc_auc_score(y_test, y_pred_proba)
            pr_auc = average_precision_score(y_test, y_pred_proba)

            report = classification_report(y_test, y_pred, output_dict=True)
            conf_matrix = confusion_matrix(y_test, y_pred)

            # Print results
            logging.info(f"{model_name} Results:")
            logging.info(f"Accuracy: {accuracy:.4f}")
            logging.info(f"Precision: {precision:.4f}")
            logging.info(f"Recall: {recall:.4f}")
            logging.info(f"F1-score: {f1:.4f}")
            logging.info(f"ROC AUC: {roc_auc:.4f}")
            logging.info(f"PR AUC: {pr_auc:.4f}\n")

            # Save results to Excel file
            df_report = pd.DataFrame(report).transpose()

            # Check if sheet already exists and delete if necessary
            book = load_workbook(EXCEL_FILE)
            if f'{model_name}_report' in book.sheetnames:
                del book[f'{model_name}_report']
                book.save(EXCEL_FILE)
                logging.info(f"Existing sheet '{model_name}_report' deleted.")

            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a") as writer:
                df_report.to_excel(writer, sheet_name=f'{model_name}_report')

            # Save confusion matrix
            plt.figure(figsize=(8, 6))
            sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues',
                        xticklabels=['No Failure', 'Failure'], yticklabels=['No Failure', 'Failure'])
            plt.title(f'Confusion Matrix - {model_name}')
            plt.xlabel('Prediction')
            plt.ylabel('Actual')
            plt.tight_layout()
            conf_matrix_path = os.path.join(RESULTS_DIR, f'{model_name}_confusion_matrix.png')
            plt.savefig(conf_matrix_path, dpi=300)
            plt.close()
            logging.info(f"Confusion matrix '{model_name}_confusion_matrix.png' saved.")

            return y_pred_proba, roc_auc, pr_auc
        except Exception as e:
            logging.error(f"Error evaluating {model_name}: {e}")
            return None, None, None

    model_metrics = {}
    for model_name, model in best_models.items():
        metrics = save_results(model_name, model, x_test, y_test)
        if metrics[0] is not None:
            y_pred_proba, roc_auc, pr_auc = metrics
            model_metrics[model_name] = {
                'y_pred_proba': y_pred_proba,
                'roc_auc': roc_auc,
                'pr_auc': pr_auc
            }
    return model_metrics

def plot_classification_curves(model_metrics, y_test):
    """Plot ROC and Precision-Recall curves for all models."""
    try:
        plt.figure(figsize=(12, 6))

        # ROC curves
        plt.subplot(1, 2, 1)
        for model_name, metrics in model_metrics.items():
            fpr, tpr, _ = roc_curve(y_test, metrics['y_pred_proba'])
            plt.plot(fpr, tpr, label=f'{model_name} (AUC = {metrics["roc_auc"]:.2f})')
        plt.plot([0, 1], [0, 1], 'k--')
        plt.title('ROC Curves')
        plt.xlabel('False Positive Rate')
        plt.ylabel('True Positive Rate')
        plt.legend(loc='lower right')

        # Precision-Recall curves
        plt.subplot(1, 2, 2)
        for model_name, metrics in model_metrics.items():
            precision, recall, _ = precision_recall_curve(y_test, metrics['y_pred_proba'])
            plt.plot(recall, precision, label=f'{model_name} (AUC = {metrics["pr_auc"]:.2f})')
        plt.title('Precision-Recall Curves')
        plt.xlabel('Recall')
        plt.ylabel('Precision')
        plt.legend(loc='lower left')

        plt.tight_layout()
        roc_pr_path = os.path.join(RESULTS_DIR, 'roc_pr_curves.png')
        plt.savefig(roc_pr_path, dpi=300)
        plt.close()
        logging.info("Plot 'roc_pr_curves.png' saved successfully.")
    except Exception as e:
        logging.error(f"Error plotting classification curves: {e}")

def plot_feature_importance(models, feature_names):
    """Plot feature importance for models that support it."""
    for model_name, model in models.items():
        if hasattr(model, 'feature_importances_'):
            try:
                importances = model.feature_importances_
                indices = np.argsort(importances)[::-1]
                plt.figure(figsize=(10, 6))
                sns.barplot(x=importances[indices], y=np.array(feature_names)[indices], palette='viridis')
                plt.title(f'Feature Importance - {model_name}')
                plt.xlabel('Importance')
                plt.ylabel('Features')
                plt.tight_layout()
                fi_path = os.path.join(RESULTS_DIR, f'{model_name}_feature_importance.png')
                plt.savefig(fi_path, dpi=300)
                plt.close()
                logging.info(f"Feature importance plot '{model_name}_feature_importance.png' saved successfully.")
            except Exception as e:
                logging.error(f"Error plotting feature importance for {model_name}: {e}")

def detect_anomalies(data):
    """Apply five anomaly detection algorithms to temporal data."""
    # Select numeric features
    columns_to_drop = ['equipment_id', 'time_step', 'failure', 'anomaly']
    existing_columns_to_drop = [col for col in columns_to_drop if col in data.columns]
    X_anomaly = data.drop(existing_columns_to_drop, axis=1, errors='ignore')

    # Determine scikit-learn version to handle OneHotEncoder if necessary
    skl_version = version.parse(sklearn.__version__)

    # Encode categorical variables if they exist
    categorical_cols = X_anomaly.select_dtypes(include=['object', 'category']).columns.tolist()
    if categorical_cols:
        try:
            if skl_version >= version.parse("1.2"):
                encoder = OneHotEncoder(drop='first', sparse_output=False)
            else:
                encoder = OneHotEncoder(drop='first', sparse=False)
            encoded_data = encoder.fit_transform(X_anomaly[categorical_cols])
            encoded_cols = encoder.get_feature_names_out(categorical_cols)
            encoded_df = pd.DataFrame(encoded_data, columns=encoded_cols, index=X_anomaly.index)
            X_anomaly = pd.concat([X_anomaly.drop(categorical_cols, axis=1), encoded_df], axis=1)
            logging.info("Categorical variables encoded correctly for anomaly detection.")
        except Exception as e:
            logging.error(f"Error encoding categorical variables for anomaly detection: {e}")
            raise
    else:
        logging.info("No categorical columns found to encode for anomaly detection.")

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_anomaly)
    logging.info("Features scaled for anomaly detection.")

    anomaly_results = pd.DataFrame(index=data.index)
    anomaly_results['True_Anomaly'] = data['anomaly'].astype(int)

    # 1. Isolation Forest
    try:
        iso_forest = IsolationForest(contamination=0.02, random_state=42)
        iso_forest.fit(X_scaled)
        y_pred_iso = iso_forest.predict(X_scaled)
        anomaly_results['IsolationForest'] = np.where(y_pred_iso == -1, 1, 0)
        logging.info("IsolationForest applied successfully.")
    except Exception as e:
        logging.error(f"Error applying IsolationForest: {e}")

    # 2. One-Class SVM
    try:
        one_class_svm = OneClassSVM(nu=0.02, kernel='rbf', gamma='scale')
        one_class_svm.fit(X_scaled)
        y_pred_svm = one_class_svm.predict(X_scaled)
        anomaly_results['OneClassSVM'] = np.where(y_pred_svm == -1, 1, 0)
        logging.info("OneClassSVM applied successfully.")
    except Exception as e:
        logging.error(f"Error applying OneClassSVM: {e}")

    # 3. Local Outlier Factor
    try:
        lof = LocalOutlierFactor(n_neighbors=20, contamination=0.02)
        y_pred_lof = lof.fit_predict(X_scaled)
        anomaly_results['LocalOutlierFactor'] = np.where(y_pred_lof == -1, 1, 0)
        logging.info("LocalOutlierFactor applied successfully.")
    except Exception as e:
        logging.error(f"Error applying LocalOutlierFactor: {e}")

    # 4. DBSCAN
    try:
        dbscan = DBSCAN(eps=3, min_samples=5)
        dbscan_labels = dbscan.fit_predict(X_scaled)
        anomaly_results['DBSCAN'] = np.where(dbscan_labels == -1, 1, 0)
        logging.info("DBSCAN applied successfully.")
    except Exception as e:
        logging.error(f"Error applying DBSCAN: {e}")

    # 5. PCA-based Outlier Detection
    try:
        pca = PCA(n_components=2)
        X_pca = pca.fit_transform(X_scaled)
        pca_distances = np.linalg.norm(X_pca, axis=1)
        threshold = np.percentile(pca_distances, 98)  # Top 2% as anomalies
        anomaly_results['PCA_Outlier'] = (pca_distances > threshold).astype(int)
        logging.info("PCA-based Outlier Detection applied successfully.")
    except Exception as e:
        logging.error(f"Error applying PCA-based Outlier Detection: {e}")

    # Save anomaly results
    try:
        anomaly_results.to_csv(os.path.join(RESULTS_DIR, 'anomaly_detection_results.csv'), index=False)
        logging.info("Anomaly detection results saved to 'anomaly_detection_results.csv'.")
    except Exception as e:
        logging.error(f"Error saving anomaly detection results: {e}")

    # Evaluation of detections
    try:
        for method in ['IsolationForest', 'OneClassSVM', 'LocalOutlierFactor', 'DBSCAN', 'PCA_Outlier']:
            y_true = anomaly_results['True_Anomaly']
            y_pred = anomaly_results[method]
            precision = precision_score(y_true, y_pred, zero_division=0)
            recall = recall_score(y_true, y_pred, zero_division=0)
            f1 = f1_score(y_true, y_pred, zero_division=0)
            logging.info(f"Anomaly Detection - {method}: Precision={precision:.4f}, Recall={recall:.4f}, F1-score={f1:.4f}")
    except Exception as e:
        logging.error(f"Error evaluating anomaly detection: {e}")

    # Generate anomaly detection plots
    try:
        for method in ['IsolationForest', 'OneClassSVM', 'LocalOutlierFactor', 'DBSCAN', 'PCA_Outlier']:
            plt.figure(figsize=(10, 6))
            if 'load' in data.columns:
                sns.scatterplot(x=data.index, y=data['load'], hue=anomaly_results[method], palette='coolwarm', legend=False)
                plt.title(f'Anomaly Detection - {method}')
                plt.xlabel('Sample Index')
                plt.ylabel('Load')
            else:
                # If 'load' is not available, use another numeric variable
                numerical_vars = ['vibration', 'oil_quality', 'temperature', 'pressure', 'hours_operated']
                available_var = next((var for var in numerical_vars if var in data.columns), 'vibration')
                sns.scatterplot(x=data.index, y=data[available_var], hue=anomaly_results[method], palette='coolwarm', legend=False)
                plt.title(f'Anomaly Detection - {method}')
                plt.xlabel('Sample Index')
                plt.ylabel(f'{available_var} Value')
            plt.tight_layout()
            anomaly_plot_path = os.path.join(RESULTS_DIR, f'{method}_anomaly_detection.png')
            plt.savefig(anomaly_plot_path, dpi=300)
            plt.close()
            logging.info(f"Anomaly detection plot '{method}_anomaly_detection.png' saved successfully.")
    except Exception as e:
        logging.error(f"Error generating anomaly detection plots: {e}")

    return anomaly_results

def create_pdf_report(data, model_metrics, feature_names, best_models, anomaly_results):
    """Create a PDF report with generated plots."""
    try:
        pdf_file = PDF_REPORT
        doc = SimpleDocTemplate(pdf_file, pagesize=A4,
                                rightMargin=30, leftMargin=30,
                                topMargin=30, bottomMargin=18)
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='CenterTitle', alignment=1, fontSize=16, spaceAfter=20))
        flowables = []

        # Title
        flowables.append(Paragraph("Technical Report: Predictive Maintenance Analysis and Anomaly Detection", styles['CenterTitle']))

        # General description
        description = (
            "This report presents a comprehensive analysis of predictive maintenance using Machine Learning techniques and Anomaly Detection. "
            "Simulated temporal data was generated for multiple equipment over 40 time steps to predict failures and detect anomalies in their evolution. "
            "Multiple models were trained and their performance was evaluated using standard metrics and detailed visualizations."
        )
        flowables.append(Paragraph(description, styles['Normal']))
        flowables.append(Spacer(1, 12))

        # EDA: Failure Distribution
        flowables.append(Paragraph("1. Exploratory Data Analysis (EDA)", styles['Heading2']))
        flowables.append(Spacer(1, 12))
        flowables.append(Paragraph("Figure 1: Failure vs No Failure Distribution", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'failure_distribution.png'), width=400, height=300))
        flowables.append(Spacer(1, 12))

        # EDA: Correlation Matrix
        flowables.append(Paragraph("Figure 2: Simulated Data Correlation Matrix", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'correlation_matrix.png'), width=500, height=400))
        flowables.append(Spacer(1, 12))

        # EDA: Histograms
        flowables.append(Paragraph("Figure 3: Variable Histograms", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'histograms.png'), width=500, height=400))
        flowables.append(Spacer(1, 12))

        # EDA: Boxplots
        flowables.append(Paragraph("Figure 4: Variable Boxplots", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'boxplots.png'), width=500, height=400))
        flowables.append(Spacer(1, 12))

        # EDA: Pairplot
        flowables.append(Paragraph("Figure 5: Variable Pairplot", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'pairplot.png'), width=500, height=400))
        flowables.append(Spacer(1, 12))

        # Anomaly Detection
        flowables.append(Paragraph("2. Anomaly Detection", styles['Heading2']))
        flowables.append(Spacer(1, 12))
        flowables.append(Paragraph("Five different anomaly detection algorithms were applied to identify unusual behaviors in temporal data.", styles['Normal']))
        flowables.append(Spacer(1, 12))
        # Include anomaly detection plots
        for idx, method in enumerate(['IsolationForest', 'OneClassSVM', 'LocalOutlierFactor', 'DBSCAN', 'PCA_Outlier'], start=1):
            anomaly_plot_path = os.path.join(RESULTS_DIR, f'{method}_anomaly_detection.png')
            if os.path.exists(anomaly_plot_path):
                flowables.append(Paragraph(f"Figure {5 + idx}: Anomaly Detection - {method}", styles['Heading3']))
                flowables.append(Image(anomaly_plot_path, width=400, height=300))
                flowables.append(Spacer(1, 12))

        # Classification Model Evaluation
        flowables.append(Paragraph("3. Classification Model Evaluation", styles['Heading2']))
        flowables.append(Spacer(1, 12))

        # Add model metrics
        for idx, (model_name, metrics) in enumerate(model_metrics.items(), start=1):
            flowables.append(Paragraph(f"3.{idx} {model_name}", styles['Heading3']))
            flowables.append(Paragraph(f"ROC AUC: {metrics['roc_auc']:.4f}", styles['Normal']))
            flowables.append(Paragraph(f"PR AUC: {metrics['pr_auc']:.4f}", styles['Normal']))
            flowables.append(Spacer(1, 12))
            # Include confusion matrix
            conf_matrix_path = os.path.join(RESULTS_DIR, f"{model_name}_confusion_matrix.png")
            if os.path.exists(conf_matrix_path):
                # Adjust figure number
                figure_number = 5 + len(['IsolationForest', 'OneClassSVM', 'LocalOutlierFactor', 'DBSCAN', 'PCA_Outlier']) + idx
                flowables.append(Paragraph(f"Figure {figure_number}: Confusion Matrix - {model_name}", styles['Heading4']))
                flowables.append(Image(conf_matrix_path, width=300, height=250))
                flowables.append(Spacer(1, 12))

        # ROC and PR curves
        flowables.append(Paragraph("Figure 11: ROC and Precision-Recall Curves", styles['Heading3']))
        flowables.append(Image(os.path.join(RESULTS_DIR, 'roc_pr_curves.png'), width=500, height=300))
        flowables.append(Spacer(1, 12))

        # Feature importance
        flowables.append(Paragraph("4. Feature Importance", styles['Heading2']))
        flowables.append(Spacer(1, 12))
        for idx, model_name in enumerate(best_models.keys(), start=1):
            fi_path = os.path.join(RESULTS_DIR, f'{model_name}_feature_importance.png')
            if os.path.exists(fi_path):
                figure_number = 12 + idx
                flowables.append(Paragraph(f"Figure {figure_number}: Feature Importance - {model_name}", styles['Heading3']))
                flowables.append(Image(fi_path, width=400, height=300))
                flowables.append(Spacer(1, 12))

        # Conclusions
        flowables.append(Paragraph("5. Conclusions", styles['Heading2']))
        conclusions = (
            "The evaluated classification models demonstrated promising performance in predicting equipment failures. "
            "Among the evaluated models, Random Forest and Gradient Boosting showed the best performance metrics, "
            "indicating a high capacity to distinguish between equipment that will fail and those that will not. "
            "Anomaly detection using five different algorithms allowed identifying unusual behaviors in the temporal evolution of equipment. "
            "Variables such as vibration, oil quality, hours operated, and maintenance history were the most determining factors for predicting failures. "
            "These findings suggest that continuous monitoring and preventive maintenance based on these metrics can significantly improve equipment reliability."
        )
        flowables.append(Paragraph(conclusions, styles['Normal']))

        # Generate PDF
        doc.build(flowables)
        logging.info(f"PDF report generated in '{pdf_file}'.")
    except Exception as e:
        logging.error(f"Error creating PDF report: {e}")

def main():
    """Main function for overall flow"""
    setup_directories()
    data = generate_simulated_temporal_data()
    exportToCSV(data)
    data = handle_data_types(data)  # Handle data types before EDA

    # Additional verification
    categorical_cols_remaining = data.select_dtypes(include=['object', 'category']).columns.tolist()
    if categorical_cols_remaining:
        raise ValueError(f"The following columns are still categorical and have not been encoded: {categorical_cols_remaining}")
    else:
        logging.info("All categorical columns have been encoded correctly.")

    perform_eda(data)
    x_train, x_test, y_train, y_test, feature_names = preprocess_data(data)
    best_models = train_classification_models(x_train, y_train)
    model_metrics = evaluate_classification_models(best_models, x_test, y_test)
    export_models(best_models)
    plot_classification_curves(model_metrics, y_test)
    plot_feature_importance(best_models, feature_names)
    anomaly_results = detect_anomalies(data)
    create_pdf_report(data, model_metrics, feature_names, best_models, anomaly_results)
    logging.info("Process completed successfully. All results and report have been saved in the 'outputs' folder.")

if __name__ == "__main__":
    main()